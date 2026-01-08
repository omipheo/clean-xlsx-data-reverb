import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import re
import os
from pathlib import Path

def is_date_value(value):
    """Check if value is a datetime object"""
    return isinstance(value, datetime)

def is_valid_price(value):
    """Check if value looks like a price"""
    if value is None:
        return False
    
    # If it's a date before 2025-07-24, it's likely a wrongly formatted price
    if is_date_value(value):
        if value < datetime(2025, 7, 24):
            return True  # This is actually a price formatted as date
        return False
    
    # Check if it's a number
    if isinstance(value, (int, float)):
        return value > 0 and value < 10000  # Reasonable price range
    
    return False

def convert_date_to_price(value):
    """Convert wrongly formatted date to price number"""
    if is_date_value(value) and value < datetime(2025, 7, 24):
        # Extract day which represents the price
        # e.g., 1900-03-20 means $20
        return value.day
    return value

def is_person_name_simple(value):
    """Simplified check if this could be a person's name"""
    if value is None or not isinstance(value, str):
        return False
    
    value = str(value).strip()
    
    # Skip empty or very short strings
    if len(value) < 5:
        return False
    
    # Skip if it ends with a period (common in pedal listings)
    if value.endswith('.'):
        return False
    
    # Skip if it contains a colon (common in pedal names like "Neunaber: Illumine")
    if ':' in value:
        return False
    
    # Skip lines with these keywords (definite non-names)
    skip_keywords = ['total', 'fmv', '$20 less', 'pedal', 'label', 'needs', 'for lot', 
                     'payout', 'boss ', 'mxr ', 'tc electronic', 'digitech ', 'ibanez ',
                     'reverb', 'delay', 'overdrive', 'distortion', 'fuzz', 'chorus',
                     'flanger', 'compressor', 'looper', 'wah', 'boost', 'tremolo',
                     'preamp', 'squeezer', 'gold', 'ridge', 'ray', 'mjolnir',
                     'vemuram', 'neunaber', 'illumine', 'xotic', 'kernom', 'mythos',
                     'eqd', 'wd orange', 'jan ray', 'acapulco', 'bb preamp']
    
    value_lower = value.lower()
    if any(keyword in value_lower for keyword in skip_keywords):
        return False
    
    # Names typically have 2-4 words
    words = value.split()
    if len(words) < 2 or len(words) > 4:
        return False
    
    # All words should start with capital letter
    if not all(word and word[0].isupper() for word in words):
        return False
    
    # Should not contain digits
    if any(c.isdigit() for c in value):
        return False
    
    # Not too long
    if len(value) > 60:
        return False
    
    # Additional check: Common first names to increase confidence
    # This is a heuristic - if the first word looks like a common first name, it's more likely to be a person
    common_first_names = ['michael', 'david', 'james', 'john', 'robert', 'william', 'richard',
                          'thomas', 'christopher', 'daniel', 'matthew', 'joseph', 'anthony',
                          'donald', 'mark', 'paul', 'steven', 'andrew', 'kenneth', 'joshua',
                          'kevin', 'brian', 'george', 'edward', 'ronald', 'timothy', 'jason',
                          'jeffrey', 'ryan', 'jacob', 'gary', 'nicholas', 'eric', 'jonathan',
                          'stephen', 'larry', 'justin', 'scott', 'brandon', 'frank', 'benjamin',
                          'gregory', 'raymond', 'samuel', 'patrick', 'alexander', 'jack', 'dennis',
                          'jerry', 'tyler', 'aaron', 'jose', 'adam', 'henry', 'nathan', 'douglas',
                          'zachary', 'peter', 'kyle', 'walter', 'ethan', 'jeremy', 'harold',
                          'keith', 'christian', 'roger', 'noah', 'gerald', 'carl', 'terry', 'sean',
                          'austin', 'arthur', 'lawrence', 'jesse', 'dylan', 'bryan', 'joe', 'jordan',
                          'billy', 'bruce', 'albert', 'willie', 'gabriel', 'logan', 'alan', 'juan',
                          'wayne', 'roy', 'ralph', 'randy', 'eugene', 'vincent', 'russell', 'elijah',
                          'louis', 'bobby', 'philip', 'johnny', 'bradley', 'neil', 'andre', 'jean',
                          'rob', 'steve', 'oliver', 'jimmy', 'martin', 'dave', 'matt', 'chris',
                          'mike', 'bill', 'bob', 'tom', 'dan', 'jim', 'tony', 'pablo', 'joey',
                          'stefan', 'salvatore', 'jimi', 'damon', 'adam', 'jake', 'oliver',
                          'vince', 'brett', 'joshua', 'jamie', 'sean', 'dennis', 'doug', 'george',
                          'william', 'tracy', 'adan', 'mitchell', 'daryl', 'guy', 'edward', 'jerry',
                          'paul', 'jean-claude', 'gandhi', 'randy', 'chris', 'mike']
    
    first_word = words[0].lower()
    
    # Strong indicator: common first name
    if first_word in common_first_names:
        return True
    
    # If not a common name, be more strict
    # Must not look like a brand or model
    brand_indicators = ['micro', 'mini', 'super', 'ultra', 'pro', 'deluxe', 'king', 'master',
                       'special', 'custom', 'classic', 'vintage', 'mk', 'v1', 'v2', 'v3', 'v4']
    if any(indicator in value_lower for indicator in brand_indicators):
        return False
    
    return False  # Default to False if no common name detected

def is_pedal_name(value):
    """Check if this looks like a pedal name"""
    if value is None or not isinstance(value, str):
        return False
    
    # Skip these rows
    skip_keywords = ['total', 'fmv', '$20 less', 'label', 'needs', 'for lot', 'payout']
    value_lower = value.lower()
    if any(keyword in value_lower for keyword in skip_keywords):
        return False
    
    # If it contains typical pedal-related words or brand names
    pedal_indicators = ['pedal', 'reverb', 'delay', 'overdrive', 'distortion', 'fuzz', 
                        'chorus', 'flanger', 'wah', 'boost', 'compressor', 'looper',
                        'boss', 'mxr', 'tc electronic', 'digitech', 'ibanez', 
                        'dunlop', 'walrus', 'eqd', 'keeley', 'jhs', 'wampler',
                        'line 6', 'behringer', 'mooer', 'pigtronix', 'earthquaker']
    
    if any(indicator in value_lower for indicator in pedal_indicators):
        return True
    
    # If it has mixed case and some length, could be a pedal
    if len(value) > 5 and any(c.isalpha() for c in value):
        return True
    
    return False

def clean_spreadsheet(input_file):
    """Clean the pedal pricing spreadsheet and return structured data"""
    
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb['Sheet1']
    
    cleaned_data = []
    current_person = None
    current_date = None
    last_valid_date = datetime(2025, 7, 24)  # Default to first date in spreadsheet
    
    # Process all rows
    for row_idx in range(1, ws.max_row + 1):
        # Get values from multiple possible columns
        col_a = ws.cell(row_idx, 1).value  # Column A
        col_b = ws.cell(row_idx, 2).value  # Column B
        col_c = ws.cell(row_idx, 3).value  # Column C
        col_d = ws.cell(row_idx, 4).value  # Column D
        col_e = ws.cell(row_idx, 5).value  # Column E
        col_f = ws.cell(row_idx, 6).value  # Column F
        col_g = ws.cell(row_idx, 7).value  # Column G
        col_h = ws.cell(row_idx, 8).value  # Column H
        
        # Check if column B has a valid date (>= 2025-07-24)
        has_valid_date = col_b and is_date_value(col_b) and col_b >= datetime(2025, 7, 24)
        
        # Check if this row might be a person's name
        is_potential_person = col_a and is_person_name_simple(col_a)
        
        # If we have a person name with a valid date, this is definitely a person
        if is_potential_person and has_valid_date:
            person_name = str(col_a).strip()
            current_person = person_name
            current_date = col_b
            last_valid_date = col_b
            continue
        
        # If we have a person name without a date, use the last valid date
        if is_potential_person and not has_valid_date:
            person_name = str(col_a).strip()
            current_person = person_name
            current_date = last_valid_date
            continue
        
        # If we don't have a current date yet, use the default
        if current_date is None:
            current_date = last_valid_date
        
        # Check for pedal in column A with price in column B
        if col_a and is_pedal_name(col_a):
            pedal_name = str(col_a).strip()
            price = None
            
            # Check column B for price
            if is_valid_price(col_b):
                price = convert_date_to_price(col_b)
                if is_date_value(price):
                    price = None
            
            # If no price in B, check column C
            if price is None and is_valid_price(col_c):
                price = convert_date_to_price(col_c)
                if is_date_value(price):
                    price = None
            
            if price is not None and isinstance(price, (int, float)):
                cleaned_data.append({
                    'person': current_person if current_person else 'Unknown',
                    'pedal_name': pedal_name,
                    'price': float(price),
                    'date': current_date
                })
        
        # Check for pedal in column D with price in column E
        if col_d and is_pedal_name(col_d):
            pedal_name = str(col_d).strip()
            price = None
            
            if is_valid_price(col_e):
                price = convert_date_to_price(col_e)
                if is_date_value(price):
                    price = None
            
            # If no price in E, check column F
            if price is None and is_valid_price(col_f):
                price = convert_date_to_price(col_f)
                if is_date_value(price):
                    price = None
            
            if price is not None and isinstance(price, (int, float)):
                cleaned_data.append({
                    'person': current_person if current_person else 'Unknown',
                    'pedal_name': pedal_name,
                    'price': float(price),
                    'date': current_date
                })
        
        # Check for pedal in column F with price in column G
        if col_f and is_pedal_name(col_f):
            pedal_name = str(col_f).strip()
            price = None
            
            if is_valid_price(col_g):
                price = convert_date_to_price(col_g)
                if is_date_value(price):
                    price = None
            
            if price is not None and isinstance(price, (int, float)):
                cleaned_data.append({
                    'person': current_person if current_person else 'Unknown',
                    'pedal_name': pedal_name,
                    'price': float(price),
                    'date': current_date
                })
        
        # Check for prices in column H (special case)
        if col_h and is_valid_price(col_h):
            price = convert_date_to_price(col_h)
            if not is_date_value(price) and isinstance(price, (int, float)):
                # Try to get pedal name from column A or D
                pedal_name = None
                if col_a and is_pedal_name(col_a):
                    pedal_name = str(col_a).strip()
                elif col_d and is_pedal_name(col_d):
                    pedal_name = str(col_d).strip()
                
                if pedal_name:
                    cleaned_data.append({
                        'person': current_person if current_person else 'Unknown',
                        'pedal_name': pedal_name,
                        'price': float(price),
                        'date': current_date
                    })
    
    wb.close()
    
    # Create DataFrame
    df = pd.DataFrame(cleaned_data)
    
    # Calculate expiration date (1 year after the date)
    df['expiration_date'] = df['date'].apply(lambda x: x + timedelta(days=365))
    
    # Format dates
    df['date'] = df['date'].dt.strftime('%Y-%m-%d')
    df['expiration_date'] = df['expiration_date'].dt.strftime('%Y-%m-%d')
    
    # Reorder columns (excluding person column - not needed in final output)
    df = df[['pedal_name', 'price', 'date', 'expiration_date']]
    
    return df

if __name__ == "__main__":
    print("Cleaning pedal pricing spreadsheet...")
    
    # Get the script directory and project root
    script_dir = Path(__file__).parent
    # project_root = script_dir.parent
    
    # Input file (in project root)
    input_file = script_dir / 'justin pricing spreadsheet.xlsx'
    
    if not input_file.exists():
        print(f"❌ Error: Input file not found at {input_file}")
        print(f"   Please ensure 'justin pricing spreadsheet.xlsx' is in the project root directory")
        exit(1)
    
    # Clean the data
    cleaned_df = clean_spreadsheet(str(input_file))
    
    print(f"\nExtracted {len(cleaned_df)} pedal entries")
    print("\nFirst 20 entries:")
    print(cleaned_df.head(20).to_string(index=False))
    
    print("\n\nLast 20 entries:")
    print(cleaned_df.tail(20).to_string(index=False))
    
    # Save to new Excel file (in project root)
    output_file = script_dir / 'cleaned_pedal_pricing.xlsx'
    cleaned_df.to_excel(str(output_file), index=False, sheet_name='Cleaned_Data')
    print(f"\n✅ Cleaned data saved to: {output_file}")
    
    # Also save as CSV for easier integration
    csv_output = script_dir / 'cleaned_pedal_pricing.csv'
    cleaned_df.to_csv(str(csv_output), index=False)
    print(f"✅ CSV version saved to: {csv_output}")
    
    # Show summary statistics
    print("\n" + "="*80)
    print("SUMMARY STATISTICS")
    print("="*80)
    print(f"Total pedals: {len(cleaned_df)}")
    print(f"Unique pedal names: {cleaned_df['pedal_name'].nunique()}")
    print(f"Date range: {cleaned_df['date'].min()} to {cleaned_df['date'].max()}")
    print(f"\nPrice statistics:")
    print(cleaned_df['price'].describe())